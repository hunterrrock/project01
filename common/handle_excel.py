"""
=======================|
Author:76              | 
Time:2021/5/4  19:01  |
=======================|
"""

import openpyxl



class HandleExcel():
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        data = list(sh.rows)
        cases = []
        title = [i.value for i in data[0]]
        for item in data[1:]:
            res = [i.value for i in item]
            dic = dict(zip(title, res))
            cases.append(dic)
        return cases

    def write_excel(self, row, column, value):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)




if __name__ == '__main__':
    excel = HandleExcel(r'D:\PythonPoject\project01\datas\wb1.xlsx', 'login')
    res = excel.read_excel()
    print(res)
    excel.write_excel(row=1, column=1, value='python')
